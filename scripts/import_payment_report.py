"""One-time migration: import ~3500 trips from another application's "detailed
payment report" export (an internal ops/payment-tracking spreadsheet, not this
app) into the `trucks` table.

Each row is one completed/in-progress trip with a real driver, vehicle, route,
and material — this comes from an internal ops system (not an anonymous mobile
contributor), so like import_legacy_field_survey.py it is inserted already
VERIFIED/PASSED and skips the telecaller review queue.

Design decisions:
  - "Trip Route" is usually one leg, "<origin> to <destination>" (e.g. "Sandila
    to Meerut"), but ~7% of rows are multi-drop trips: several comma-separated
    legs, e.g. "Jaipur to Khagaria, Jaipur to Saharsa" (sometimes with a
    duplicate leg, e.g. "A to B, A to B"). A broker searching ANY place on the
    route should find this truck, so each trip becomes one row per distinct
    place mentioned (deduped) — everything else identical across rows. Each is
    set on `city` (not just `location`) so it also populates the Location
    filter dropdown on `/`, not just free-text search.
  - source = image_api, verification_status=VERIFIED, review_status=PASSED,
    processing_status=DONE — same rationale as import_legacy_field_survey.py.
  - phone_number = Driver Phone only (the callable contact this data is meant
    to provide). Vehicle Owner Phone is folded into other_text for reference,
    not treated as a second callable number.
  - company_name = Vehicle Owner Name (fleet/owner name — the closest analog to
    what OCR would read off the truck for video/stream sightings).
  - material_type = Material column, directly. driver_name = Driver Name.
  - This export is overwhelmingly payment/invoice bookkeeping (Consignor,
    Bilty Number, Contractor, weight, dozens of financial columns) that has no
    bearing on a broker calling a truck — skipped, except a few folded into
    other_text for reference (Consignor, Weight, Trip Stage, Bilty Number,
    Vehicle Owner Phone).
  - detected_at = Trip Date ("DD/MM/YYYY hh:mm AM/PM"). One batch of ~11 rows
    has a corrupted year (e.g. "08/10/0203" instead of "...2023" — an export
    artifact); when the parsed year is implausible (<2000 or >current year+1)
    this falls back to Bill Creation Date (a reliable ISO timestamp), then to
    import time.
  - Skipped: the trailing "Total" summary row, and any row missing a usable
    vehicle plate, a valid 10-digit driver phone, or a parseable Trip Route.
  - Idempotent: source_ref = "payment_report_row_<N>_<leg index>" (N = the
    sheet row number). Re-running skips any row whose N already has rows
    imported, so it's safe to run again (e.g. once locally, then again on prod).

Usage:
    .venv\\Scripts\\python.exe scripts\\import_payment_report.py --dry-run
    .venv\\Scripts\\python.exe scripts\\import_payment_report.py
    .venv\\Scripts\\python.exe scripts\\import_payment_report.py --xlsx path\\to\\file.xlsx
"""
import argparse
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402
from pipeline.db import SourceType, Truck, get_session_factory, init_db, database_url  # noqa: E402

DEFAULT_XLSX = str(Path(__file__).resolve().parent.parent / "detailed-payment-report.xlsx")
SHEET_NAME = "Consolidated Workbook"

IST = timezone(timedelta(hours=5, minutes=30), name="IST")
SOURCE_REF_PREFIX = "payment_report_row_"
COMMIT_EVERY = 500  # chunk commits so one giant transaction doesn't strain the small prod VM

_MOBILE_RE = re.compile(r'^[6-9]\d{9}$')
_ROUTE_RE = re.compile(r'^(.+?)\s+to\s+(.+)$', re.IGNORECASE)
_TRIP_DATE_FMT = "%d/%m/%Y %I:%M %p"


def _clean_phone(raw):
    if not raw:
        return None
    digits = re.sub(r'\D', '', str(raw))
    digits = re.sub(r'^(0091|91|0)(?=\d{10}$)', '', digits)
    return digits if _MOBILE_RE.match(digits) else None


def _clean_plate(raw):
    v = re.sub(r'[^A-Za-z0-9]', '', str(raw or '')).upper()
    return v or None


def _extract_locations(raw):
    """Trip Route is usually one leg ('A to B') but can be several comma-
    separated legs for a multi-drop trip ('A to B, A to C'), sometimes with
    duplicate legs ('A to B, A to B'). Returns the deduped, ordered list of
    every distinct place mentioned across all legs (e.g. ['A', 'B', 'C'])."""
    if not raw:
        return []
    places, seen = [], set()
    for segment in str(raw).split(","):
        m = _ROUTE_RE.match(segment.strip())
        if not m:
            continue
        for place in (m.group(1).strip(), m.group(2).strip()):
            key = place.lower()
            if place and key not in seen:
                seen.add(key)
                places.append(place)
    return places


def _parse_trip_date_raw(raw):
    try:
        return datetime.strptime(str(raw).strip(), _TRIP_DATE_FMT).replace(tzinfo=IST)
    except (ValueError, AttributeError):
        return None


def _parse_iso(raw):
    try:
        return datetime.fromisoformat(str(raw).strip())
    except (ValueError, AttributeError, TypeError):
        return None


def _parse_trip_date(trip_date_raw, bill_creation_raw):
    dt = _parse_trip_date_raw(trip_date_raw)
    current_year = datetime.now().year
    if dt and 2000 <= dt.year <= current_year + 1:
        return dt
    fallback = _parse_iso(bill_creation_raw)
    if fallback:
        return fallback
    return dt or datetime.now(IST)


def _clean_material(raw):
    """A handful of rows repeat the same material several times, comma-
    separated (e.g. 'Earthing material, Earthing material, ...' x6, mirroring
    the duplicate-leg pattern in Trip Route) — long enough to overflow the
    64-char column. Dedupe (case-insensitive, order-preserving) and hard-cap
    as a safety net for anything still too long after that."""
    if not raw:
        return None
    seen, parts = set(), []
    for tok in str(raw).split(","):
        tok = tok.strip()
        key = tok.lower()
        if tok and key not in seen:
            seen.add(key)
            parts.append(tok)
    cleaned = ", ".join(parts)
    return cleaned[:64] if cleaned else None


def _compose_other_text(row):
    parts = []
    if row.get("Consignor"):
        parts.append(f"Consignor: {row['Consignor']}")
    if row.get("Vehicle Owner Phone"):
        parts.append(f"Owner phone: {row['Vehicle Owner Phone']}")
    if row.get("Weight (Ton)"):
        parts.append(f"Weight: {row['Weight (Ton)']}T")
    if row.get("Trip Stage"):
        parts.append(f"Trip stage: {row['Trip Stage']}")
    if row.get("Bilty Number"):
        parts.append(f"Bilty #: {row['Bilty Number']}")
    parts.append(f"Migrated from payment report (sheet row {row['_rownum']})")
    return "\n".join(parts)


def _build_trip(row):
    """Returns (list of Truck [one per distinct place on the route], None) to
    insert, or (None, skip_reason)."""
    plate = _clean_plate(row.get("Booked Vehicle"))
    if not plate:
        return None, "no vehicle plate"
    phone = _clean_phone(row.get("Driver Phone"))
    if not phone:
        return None, "no valid 10-digit driver phone"
    places = _extract_locations(row.get("Trip Route"))
    if not places:
        return None, f"unparseable Trip Route {row.get('Trip Route')!r}"

    detected_at = _parse_trip_date(row.get("Trip Date"), row.get("Bill Creation Date"))
    company_name = (str(row.get("Vehicle Owner Name") or "")).strip() or None
    driver_name = (str(row.get("Driver Name") or "")).strip() or None
    material_type = _clean_material(row.get("Material"))
    other_text = _compose_other_text(row)

    trucks = []
    for i, place in enumerate(places):
        trucks.append(Truck(
            detected_at=detected_at,
            source=SourceType.image_api,
            source_ref=f"{SOURCE_REF_PREFIX}{row['_rownum']}_{i}",
            license_plate=plate,
            plate_confidence=None,
            company_name=company_name,
            phone_number=phone,
            phone_reported=phone,
            vehicle_type="TRUCK",
            city=place,
            location=place,
            material_type=material_type,
            other_text=other_text,
            driver_name=driver_name,
            reported_by="Payment report import",
            verification_status="VERIFIED",
            review_status="PASSED",
            reviewed_by="Payment report import",
            reviewed_at=datetime.now(IST),
            review_note="Migrated from another application's payment/trip report",
            processing_status="DONE",
            processed_at=detected_at,
            image_keys=None,
        ))
    return trucks, None


def _load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(header, raw))
        row["_rownum"] = i
        rows.append(row)
    return rows


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to the .xlsx export")
    p.add_argument("--dry-run", action="store_true", help="Parse and report only; no DB writes")
    args = p.parse_args()

    print(f"Loading: {args.xlsx}")
    rows = _load_rows(args.xlsx)
    print(f"Parsed {len(rows)} data rows.")

    print(f"Target DB: {database_url()}")
    init_db()
    Session = get_session_factory()
    s = Session()
    try:
        existing_refs = set(s.execute(
            select(Truck.source_ref).where(Truck.source_ref.like(f"{SOURCE_REF_PREFIX}%"))
        ).scalars())
        # A trip's rows are "<prefix><rownum>_<leg index>" — recover the already-
        # imported row numbers regardless of how many legs each trip produced.
        existing_rownums = {
            m.group(1) for m in (re.match(rf"^{SOURCE_REF_PREFIX}(\d+)_\d+$", r) for r in existing_refs) if m
        }

        to_insert, skipped, already = [], [], []
        for row in rows:
            if str(row["_rownum"]) in existing_rownums:
                already.append(row["_rownum"])
                continue
            trucks, reason = _build_trip(row)
            if trucks is None:
                skipped.append((row["_rownum"], row.get("Booked Vehicle") or "(no plate)", reason))
            else:
                to_insert.append((row, trucks))

        total_rows_to_insert = sum(len(t) for _, t in to_insert)
        print(f"\nTrips to import: {len(to_insert)}  ({total_rows_to_insert} DB rows, 2 per trip)")
        print(f"Already imported (source_ref match): {len(already)}")
        print(f"Skipped: {len(skipped)}")
        if skipped:
            print("\nSkipped rows (review these manually if needed):")
            for rownum, plate, reason in skipped[:30]:
                print(f"  row {rownum:>5}  {plate:<14}  {reason}")
            if len(skipped) > 30:
                print(f"  ... and {len(skipped) - 30} more")

        if args.dry_run:
            print("\n--dry-run: no changes written.")
            for row, trucks in to_insert[:5]:
                for t in trucks:
                    print(f"  [preview] {t.source_ref}  plate={t.license_plate}  phone={t.phone_number}  "
                          f"company={t.company_name}  city={t.city}  material={t.material_type}  "
                          f"detected_at={t.detected_at}")
            if len(to_insert) > 5:
                print(f"  ... and {len(to_insert) - 5} more trips")
            return

        inserted = 0
        for i, (row, trucks) in enumerate(to_insert, start=1):
            for t in trucks:
                s.add(t)
                inserted += 1
            if inserted % COMMIT_EVERY < 2:  # commit roughly every COMMIT_EVERY rows
                s.commit()
            if i % 200 == 0:
                print(f"  ... {i}/{len(to_insert)} trips processed")
        s.commit()
        print(f"\nOK — inserted {inserted} rows ({len(to_insert)} trips).")
    finally:
        s.close()


if __name__ == "__main__":
    main()
